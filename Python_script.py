import os
import openpyxl
from jinja2 import Environment, FileSystemLoader
import json

input_workbook = openpyxl.load_workbook('ACI-script-CaaS Nokia_Ericsson CNF project.xlsm', data_only=True)
sheets_objects = input_workbook.worksheets

# Script to create Tenant
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("Tenant.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'Tenant':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            Tenant = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            description = str(row[1].value) if row[1].value else ''
            site = str(row[2].value) if row[2].value else ''
			
            Tenant['name'] = name
            Tenant['description'] = description
            Tenant['site'] = site
            json_file_name = f"{name}.json"
            json_data = generate_jinja_template(Tenant, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create Contract
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("Contract.j2")
    content = template.render(items=values_dict)
    with open(file_name, 'w') as file:
        file.write(content)

execute_bat_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'execute.bat')
with open(execute_bat_path, 'w') as bat_file:
    for tab in sheets_objects:
        if str(tab.title) == 'contract':
            for i, row in enumerate(tab.iter_rows(min_row=2)):
                Contract = {}
                if row[0].value:
                    contract_name = str(row[0].value)
                else:
                    continue
                subject = str(row[1].value) if row[1].value else ''
                description = str(row[2].value) if row[2].value else ''
                tenant = str(row[5].value) if row[5].value else ''
                scope = str(row[6].value) if row[6].value else ''
                Service_Graph_template_name = str(row[14].value) if row[14].value else ''
                site = str(row[15].value) if row[15].value else ''

                Contract['contract_name'] = contract_name
                Contract['description'] = description
                Contract['subject'] = subject
                Contract['tenant'] = tenant
                Contract['scope'] = scope
                Contract['Service_Graph_template_name'] = Service_Graph_template_name
                Contract['site'] = site			

                json_file_name = f"{contract_name}.json"
                json_data = generate_jinja_template(Contract, json_file_name)
                if site == "Meta":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
                elif site == "Pallini":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
                elif site == "KAL":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')



# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("VRF.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'vrf':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            VRF = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            description = str(row[1].value) if row[1].value else ''
            tenant = str(row[2].value) if row[2].value else ''
            contract_name = str(row[3].value) if row[3].value else ''
            site = str(row[4].value) if row[4].value else ''
			
            VRF['name'] = name
            VRF['description'] = description
            VRF['tenant'] = tenant
            VRF['contract_name'] = contract_name
            VRF['site'] = site
            json_file_name = f"{name}.json"
            json_data = generate_jinja_template(VRF, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create L3out
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("L3out.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'l3out':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            l3out = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            tenant = str(row[1].value) if row[1].value else ''
            vrf = str(row[2].value) if row[2].value else ''
            enable_bgp = str(row[3].value) if row[3].value else ''
            enable_ospf = str(row[4].value) if row[4].value else ''
            l3out_domain = str(row[7].value) if row[7].value else ''
            site = str(row[10].value) if row[10].value else ''
			
            l3out['name'] = name
            l3out['tenant'] = tenant
            l3out['vrf'] = vrf
            l3out['l3out_domain'] = l3out_domain
            l3out['enable_bgp'] = enable_bgp
            l3out['enable_ospf'] = enable_ospf
            l3out['site'] = site
            json_file_name = f"{name}_{i}.json"
            json_data = generate_jinja_template(l3out, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create L3out Node Profile
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("L3out_Node_Profile.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'l3out-node-profile':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            L3out_Node_Profile_dict = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            l3out = str(row[1].value) if row[1].value else ''
            tenant = str(row[2].value) if row[2].value else ''
            node_id = str(row[3].value) if row[3].value else ''
            router_id = str(row[5].value) if row[5].value else ''
            router_id_as_loopback = str(row[6].value) if row[6].value else ''
            bgp = str(row[7].value) if row[7].value else ''
            Loopback = str(row[9].value) if row[9].value else ''
            BGP_profile_timers = str(row[10].value) if row[10].value else ''
            site = str(row[11].value) if row[11].value else ''
			
            L3out_Node_Profile_dict['name'] = name
            L3out_Node_Profile_dict['l3out'] = l3out
            L3out_Node_Profile_dict['tenant'] = tenant
            L3out_Node_Profile_dict['node_id'] = node_id
            L3out_Node_Profile_dict['router_id'] = router_id
            L3out_Node_Profile_dict['router_id_as_loopback'] = router_id_as_loopback
            L3out_Node_Profile_dict['bgp'] = bgp
            L3out_Node_Profile_dict['Loopback'] = Loopback
            L3out_Node_Profile_dict['BGP_profile_timers'] = BGP_profile_timers
            L3out_Node_Profile_dict['site'] = site
			
            json_file_name = f"{name}_{l3out}_NodeProfile_{i}.json"
            json_data = generate_jinja_template(L3out_Node_Profile_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create L3out Static Routes
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, next_hops, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("Static Routes.j2")
    content = template.render(items=values_dict, nexthops=next_hops)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data

for tab in sheets_objects:
    if str(tab.title) == 'static-routes':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            Static_Routes = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            tenant = str(row[1].value) if row[1].value else ''
            l3out = str(row[2].value) if row[2].value else ''
            node_id = str(row[3].value) if row[3].value else ''
            aggregate = str(row[4].value) if row[4].value else ''
            preference = str(row[5].value) if row[5].value else ''
            site = str(row[6].value) if row[6].value else ''
            static_IP = str(row[7].value) if row[7].value else ''
            next_hops = []
            max_col_index = len(row) - 1  # Calculate the maximum column index
            for j in range(8, max_col_index + 1):  # Adjust the loop range
                nexthop = str(row[j].value) if row[j].value else ''
                if nexthop:
                    next_hops.append(nexthop)
            
            Static_Routes['name'] = name
            Static_Routes['tenant'] = tenant
            Static_Routes['l3out'] = l3out
            Static_Routes['node_id'] = node_id
            Static_Routes['aggregate'] = aggregate
            Static_Routes['preference'] = preference
            Static_Routes['site'] = site
            Static_Routes['static_IP'] = static_IP
            Static_Routes['nexthops'] = next_hops
			
            json_file_name = f"{name}_StaticRoute_{i}.json"
            json_data = generate_jinja_template(Static_Routes, next_hops, json_file_name)
            with open("execute.bat", 'a') as bat_file:
                if site == "Meta":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
                elif site == "Pallini":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
                elif site == "KAL":
                    bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')
					
# Script to create L3out Interface Profile
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("L3out_Int_profile.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'l3out-int-profile':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            l3out_int_profile_dict = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            l3out_node_profile = str(row[1].value) if row[1].value else ''
            l3out = str(row[2].value) if row[2].value else ''
            tenant = str(row[3].value) if row[3].value else ''
            interface_type = str(row[4].value) if row[4].value else ''
            path_type = str(row[5].value) if row[5].value else ''
            vlan_encap_id = str(row[6].value) if row[6].value else ''
            trunk_mode = str(row[7].value) if row[7].value else ''
            left_node_id = str(row[8].value) if row[8].value else ''
            right_node_id = str(row[9].value) if row[9].value else ''
            int_pol_group = str(row[11].value) if row[11].value else ''			
            ip_addr_side_a = str(row[13].value) if row[13].value else ''
            ip_addr_side_b = str(row[14].value) if row[14].value else ''	
            secondary_ip = str(row[15].value) if row[15].value else ''
            site = str(row[16].value) if row[16].value else ''
            l3out_int_profile_dict['name'] = name
            l3out_int_profile_dict['l3out_node_profile'] = l3out_node_profile
            l3out_int_profile_dict['l3out'] = l3out
            l3out_int_profile_dict['tenant'] = tenant
            l3out_int_profile_dict['interface_type'] = interface_type
            l3out_int_profile_dict['path_type'] = path_type
            l3out_int_profile_dict['vlan_encap_id'] = vlan_encap_id
            l3out_int_profile_dict['trunk_mode'] = trunk_mode
            l3out_int_profile_dict['left_node_id'] = left_node_id
            l3out_int_profile_dict['right_node_id'] = right_node_id
            l3out_int_profile_dict['int_pol_group'] = int_pol_group			
            l3out_int_profile_dict['ip_addr_side_a'] = ip_addr_side_a
            l3out_int_profile_dict['ip_addr_side_b'] = ip_addr_side_b
            l3out_int_profile_dict['secondary_ip'] = secondary_ip
            l3out_int_profile_dict['site'] = site
            json_file_name = f"{name}_{int_pol_group}_InterfaceProfile_{i}.json"
            json_data = generate_jinja_template(l3out_int_profile_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create BGP
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("ACI_BGP.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
	
for tab in sheets_objects:
    if str(tab.title) == 'BGP':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            BGP_dict = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            l3out_node_profile = str(row[1].value) if row[1].value else ''
            l3out = str(row[2].value) if row[2].value else ''
            tenant = str(row[3].value) if row[3].value else ''
            left_node_id = str(row[4].value) if row[4].value else ''
            right_node_id = str(row[5].value) if row[5].value else ''
            interface_policy_group = str(row[6].value) if row[6].value else ''
            bgp_peer_ip = str(row[7].value) if row[7].value else ''
            remote_bgp_as = str(row[9].value) if row[9].value else ''
            ttl = str(row[10].value) if row[10].value else ''
            status = str(row[12].value) if row[12].value else ''
            site = str(row[13].value) if row[13].value else ''
			
            BGP_dict['name'] = name
            BGP_dict['bgp_peer_ip'] = bgp_peer_ip
            BGP_dict['tenant'] = tenant
            BGP_dict['interface_policy_group'] = interface_policy_group
            BGP_dict['left_node_id'] = left_node_id
            BGP_dict['right_node_id'] = right_node_id
            BGP_dict['l3out'] = l3out
            BGP_dict['l3out_node_profile'] = l3out_node_profile
            BGP_dict['ttl'] = ttl
            BGP_dict['remote_bgp_as'] = remote_bgp_as
            BGP_dict['status'] = status
            BGP_dict['site'] = site
			
            json_file_name = f"{interface_policy_group}_{l3out}_BGP_{i}.json"
            json_data = generate_jinja_template(BGP_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create L3out External EPG and Relative subnets
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("L3out_External_EPG.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'external-epg':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            L3out_External_EPG = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            l3out = str(row[1].value) if row[1].value else ''
            tenant = str(row[2].value) if row[2].value else ''
            subnet = str(row[3].value) if row[3].value else ''
            consumed_contract = str(row[4].value) if row[4].value else ''
            site = str(row[5].value) if row[5].value else ''
			
            L3out_External_EPG['name'] = name
            L3out_External_EPG['l3out'] = l3out
            L3out_External_EPG['tenant'] = tenant
            L3out_External_EPG['subnet'] = subnet
            L3out_External_EPG['consumed_contract'] = consumed_contract
            L3out_External_EPG['site'] = site
			
            json_file_name = f"{name}_{l3out}_{i}.json"
            json_data = generate_jinja_template(L3out_External_EPG, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create Rule of Route MAP		
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("Create_Rule_Route_MAP.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'Create-Rule of Route MAP':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            Create_Rule_Route_MAP = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            tenant = str(row[1].value) if row[1].value else ''
            subnet = str(row[2].value) if row[2].value else ''
            aggregate = str(row[3].value) if row[3].value else ''
            site = str(row[4].value) if row[4].value else ''
			
            Create_Rule_Route_MAP['name'] = name
            Create_Rule_Route_MAP['tenant'] = tenant
            Create_Rule_Route_MAP['subnet'] = subnet
            Create_Rule_Route_MAP['aggregate'] = aggregate
            Create_Rule_Route_MAP['site'] = site
			
            json_file_name = f"{name}_{i}.json"
            json_data = generate_jinja_template(Create_Rule_Route_MAP, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create a Route MAP
# Modified function to take in file path instead of just file name
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("Route_MAP_Creation.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'Route-MAP Creation':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            Create_Route_MAP = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            tenant = str(row[1].value) if row[1].value else ''
            l3out = str(row[2].value) if row[2].value else ''
            Route_Context = str(row[3].value) if row[3].value else ''
            rule = str(row[4].value) if row[4].value else ''
            site = str(row[5].value) if row[5].value else ''
			
            Create_Route_MAP['name'] = name
            Create_Route_MAP['tenant'] = tenant
            Create_Route_MAP['l3out'] = l3out
            Create_Route_MAP['Route_Context'] = Route_Context
            Create_Route_MAP['rule'] = rule
            Create_Route_MAP['site'] = site
			
            json_file_name = f"{name}_{i}.json"
            json_data = generate_jinja_template(Create_Route_MAP, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

#Script to create Bridge domains
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("BD.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'bridge-domain':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            BD_dict = {}
            if row[0].value:
                tenant = str(row[0].value)
                if row[5].value == 'L2':
                    bd_gateway_ip = '' 
                elif row[5].value == 'L3':
                    bd_gateway_ip = str(row[6].value)
                else:
                    continue
            else:
                continue
            context = str(row[1].value) if row[1].value else ''
            name = str(row[2].value) if row[2].value else ''
            description = str(row[4].value) if row[4].value else ''
            bd_type = str(row[5].value) if row[5].value else ''    
            l3_out = str(row[7].value) if row[7].value else ''
            subnet_type = str(row[8].value) if row[8].value else ''
            l2_unknown_unicast = str(row[9].value) if row[9].value else ''
            l3_unknown_multicast = str(row[10].value) if row[10].value else ''
            multi_dest_flood = str(row[11].value) if row[11].value else ''
            arp_flood = str(row[12].value) if row[12].value else ''
            unicast_routing = str(row[13].value) if row[13].value else ''
            limit_ip_learning_to_subnet = str(row[14].value) if row[14].value else ''
            site = str(row[18].value) if row[18].value else ''			
            BD_dict['tenant'] = tenant
            BD_dict['context'] = context
            BD_dict['name'] = name
            BD_dict['description'] = description
            BD_dict['bd_type'] = bd_type
            BD_dict['bd_gateway_ip'] = bd_gateway_ip
            BD_dict['l3_out'] = l3_out
            BD_dict['subnet_type'] = subnet_type
            BD_dict['l2_unknown_unicast'] = l2_unknown_unicast
            BD_dict['l3_unknown_multicast'] = l3_unknown_multicast
            BD_dict['multi_dest_flood'] = multi_dest_flood
            BD_dict['arp_flood'] = arp_flood
            BD_dict['unicast_routing'] = unicast_routing
            BD_dict['limit_ip_learning_to_subnet'] = limit_ip_learning_to_subnet
            BD_dict['site'] = site
            json_file_name = f"{name}.json"
            json_data = generate_jinja_template(BD_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')

# Script to create Application EPGs
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("EPG.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'end-point-group':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            EPG_dict = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            description = str(row[2].value) if row[2].value else ''
            tenant = str(row[3].value) if row[3].value else ''
            app_profile = str(row[4].value) if row[4].value else ''
            phys_domain = str(row[7].value) if row[7].value else ''
            default_contract = str(row[10].value) if row[10].value else ''
            site = str(row[11].value) if row[11].value else ''
            bridge_domain = str(row[5].value) if row[5].value else ''
            EPG_dict['name'] = name
            EPG_dict['description'] = description
            EPG_dict['app_profile'] = app_profile
            EPG_dict['tenant'] = tenant
            EPG_dict['phys_domain'] = phys_domain
            EPG_dict['default_contract'] = default_contract
            EPG_dict['bridge_domain'] = bridge_domain
            EPG_dict['site'] = site
            json_file_name = f"{name}_{i}.json"
            json_data = generate_jinja_template(EPG_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
             if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
             elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
             elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')


# Script to deploy Policy Group on Application EPGs
def generate_jinja_template(values_dict, file_name):
    environment = Environment(loader=FileSystemLoader("templates/"))
    template = environment.get_template("EPG_physical_path.j2")
    content = template.render(items=values_dict)
    json_data = json.dumps(json.loads(content), indent=4)
    with open(file_name, 'w') as file:
        file.write(content)
    return json_data
    
for tab in sheets_objects:
    if str(tab.title) == 'epg-static-binding':
        for i, row in enumerate(tab.iter_rows(min_row=2)):
            EPG_dict = {}
            if row[0].value:
                name = str(row[0].value)
            else:
                continue
            tenant = str(row[2].value) if row[2].value else ''
            app_profile = str(row[1].value) if row[1].value else ''
            interface_policy_group = str(row[4].value) if row[4].value else ''
            left_node_id = str(row[6].value) if row[6].value else ''
            right_node_id = str(row[7].value) if row[7].value else ''
            encap_vlan_id = str(row[9].value) if row[9].value else ''
            site = str(row[12].value) if row[12].value else ''
            EPG_dict['name'] = name
            EPG_dict['app_profile'] = app_profile
            EPG_dict['tenant'] = tenant
            EPG_dict['interface_policy_group'] = interface_policy_group
            EPG_dict['left_node_id'] = left_node_id
            EPG_dict['right_node_id'] = right_node_id
            EPG_dict['encap_vlan_id'] = encap_vlan_id
            EPG_dict['site'] = site
            json_file_name = f"{name}_{interface_policy_group}_ApplicationEPG_{i}.json"
            json_data = generate_jinja_template(EPG_dict, json_file_name)
            with open("execute.bat", 'a') as bat_file:
              if site == "Meta":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.43.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "Pallini":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.40.145.1/api/mo/uni.json\nPAUSE\n')
              elif site == "KAL":
                bat_file.write(f'@ECHO OFF\ncurl -k -b COOKIE.txt -X POST --data "@{json_file_name}" -H "Content-Type: application/json" https://10.54.145.1/api/mo/uni.json\nPAUSE\n')
				